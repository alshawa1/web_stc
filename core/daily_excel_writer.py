import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


# ═══════════════════════════════════════════════════════════════════════════
#  EXECUTIVE COMFORTABLE PALETTE (مريح للعين وواضح للمدير)
# ═══════════════════════════════════════════════════════════════════════════
# خلفيات هادئة وراقية
NAVY_PRIMARY  = "1E3A8A"   # أزرق ملكي كلاسيكي هادئ (البانر وعناوين الأقسام)
BLUE_HEADER   = "2563EB"   # أزرق تنفيذي مميز لرؤوس الجداول
CARD_BG       = "F0F7FF"   # خلفية كروت المؤشرات: سماوي فاتح مريح جداً
CARD_BORDER   = "BFDBFE"   # إطار كروت ناعم
ALT_ROW_BG    = "F8FAFC"   # خلفية صفوف متبادلة: رمادي ثلجي ناعم جداً
WHITE_BG      = "FFFFFF"   # صفوف بيضاء نقية
TOTAL_BG      = "DBEAFE"   # صف الإجمالي: أزرق هادئ مميز

# نصوص واضحة ومتباينة للقراءة المريحة (بدون كتابة بيضاء على أبيض!)
TEXT_DARK     = "0F172A"   # أسود كحلي غامق واضح جداً لكافة البيانات والأرقام
TEXT_MUTED    = "475569"   # رمادي داكن للعناوين الفرعية والمسميات
TEXT_WHITE    = "FFFFFF"   # أبيض ناصع فقط على خلفيات زرقاء داكنة
TEXT_NAVY     = "1E3A8A"   # كحلي عريض للإجماليات والقيم البارزة

# الخطوط
FONT_NAME = "Segoe UI"
BANNER_FONT   = Font(name=FONT_NAME, size=15, bold=True, color=TEXT_WHITE)
SEC_FONT      = Font(name=FONT_NAME, size=11, bold=True, color=TEXT_WHITE)
HDR_FONT      = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_WHITE)
DATA_FONT     = Font(name=FONT_NAME, size=10, bold=False, color=TEXT_DARK)
TOT_FONT      = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_NAVY)
KPI_VAL_FONT  = Font(name=FONT_NAME, size=13, bold=True, color=TEXT_NAVY)
KPI_LBL_FONT  = Font(name=FONT_NAME, size=9, bold=True, color=TEXT_MUTED)

# الإطارات
BORDER_THIN   = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
BORDER_TOTAL  = Border(
    left=Side(style='thin', color='93C5FD'),
    right=Side(style='thin', color='93C5FD'),
    top=Side(style='medium', color='1E3A8A'),
    bottom=Side(style='double', color='1E3A8A')
)
BORDER_KPI    = Border(
    left=Side(style='medium', color=CARD_BORDER),
    right=Side(style='medium', color=CARD_BORDER),
    top=Side(style='medium', color=CARD_BORDER),
    bottom=Side(style='medium', color=CARD_BORDER)
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT  = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP   = Alignment(horizontal='center', vertical='center', wrap_text=True)

FILL_BANNER  = PatternFill("solid", fgColor=NAVY_PRIMARY)
FILL_SEC     = PatternFill("solid", fgColor=NAVY_PRIMARY)
FILL_HDR     = PatternFill("solid", fgColor=BLUE_HEADER)
FILL_HDR_VIN = PatternFill("solid", fgColor="1E40AF")
FILL_CARD    = PatternFill("solid", fgColor=CARD_BG)
FILL_ALT     = PatternFill("solid", fgColor=ALT_ROW_BG)
FILL_WHITE   = PatternFill("solid", fgColor=WHITE_BG)
FILL_TOTAL   = PatternFill("solid", fgColor=TOTAL_BG)


def auto_width(ws, max_col=25, min_w=10, max_w=28):
    for col in ws.columns:
        if col[0].column > max_col:
            continue
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[letter].width = min(max(max_len + 3, min_w), max_w)


def write_header_row(ws, row, headers, fill=FILL_HDR, height=22):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[row].height = height


def write_data_row(ws, row_num, row_data, headers, alt=False):
    is_tot = str(row_data.iloc[0]).startswith("📊")
    for i, val in enumerate(row_data, 1):
        c_name = headers[i - 1]
        fmt = None
        if pd.isna(val) or str(val).strip().lower() in ('nan', 'none', 'null'):
            if i == 1:
                val = 'غير محدد'
            elif '%' in c_name:
                val = 0.0
                fmt = "0.0%"
            elif any(k in c_name for k in ["مديونية", "تحصيل"]):
                val = 0.0
                fmt = '#,##0 "﷼"'
            else:
                val = 0
                fmt = "#,##0"
        elif isinstance(val, (int, float, np.integer, np.floating)) and not np.isnan(float(val)):
            fval = float(val)
            if "%" in c_name:
                val = fval / 100.0
                fmt = "0.0%"
            elif any(k in c_name for k in ["مديونية", "تحصيل"]):
                val = fval
                fmt = '#,##0 "﷼"'
            else:
                val = int(fval) if fval == int(fval) else fval
                fmt = "#,##0"
        else:
            val = str(val).strip()
            if val.lower() in ('nan', 'none', 'null', ''):
                val = 'غير محدد' if i == 1 else ''

        cell = ws.cell(row=row_num, column=i, value=val)
        if is_tot:
            cell.fill = FILL_TOTAL
            cell.font = TOT_FONT
            cell.border = BORDER_TOTAL
        else:
            cell.fill = FILL_ALT if alt else FILL_WHITE
            cell.font = DATA_FONT
            cell.border = BORDER_THIN

        cell.alignment = ALIGN_RIGHT if i == 1 else ALIGN_CENTER
        if fmt:
            cell.number_format = fmt


def section_title(ws, row, text, n_cols=18):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = FILL_SEC
    cell.font = SEC_FONT
    cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    ws.row_dimensions[row].height = 24


def add_styled_bar_chart(ws, anchor, title, data_ref, cats_ref, width=17, height=10, orient="bar"):
    ch = BarChart()
    ch.type = orient
    ch.style = 10
    ch.title = title
    ch.width = width
    ch.height = height
    ch.legend = None
    ch.dataLabels = DataLabelList()
    ch.dataLabels.showVal = True
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ws.add_chart(ch, anchor)
    return ch


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN GENERATOR
# ═══════════════════════════════════════════════════════════════════════════
def generate_styled_daily_excel(port_table, sup_df, col_df, df_pay_filtered,
                                report_date, contact_table=None, vintage_table=None):
    """
    Generates an executive, eye-friendly Excel Report:
    - Clean, readable corporate light palette (No white-on-white text!).
    - Two comprehensive tables:
      1. Main Performance & Contact Rates Table (المحافظ / تاريخ الإسناد).
      2. Debt Vintage Table (عمر الدين وسنة فصل الخدمة).
    - Clear, well-arranged professional charts next to tables.
    - Payments details on a separate tab.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير التنفيذي"
    ws.views.sheetView[0].rightToLeft = True
    ws.views.sheetView[0].showGridLines = True

    first_col = port_table.columns[0] if port_table is not None and not port_table.empty else "المحفظة"
    n_cols_main = len(port_table.columns) if port_table is not None else 13
    TOTAL_SPAN = max(n_cols_main, 14)

    # Freeze under header
    ws.freeze_panes = "A7"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 1. BANNER
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_SPAN)
    bc = ws.cell(row=1, column=1, value=f"📈  التقرير اليومي التنفيذي — فولو اب  |  {report_date}  |  نظام مهاره للتحصيل")
    bc.fill = FILL_BANNER
    bc.font = BANNER_FONT
    bc.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 36

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 2. KPI CARDS (مريحة للعين وخلفية ناعمة وقيم كحلية واضحة)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    def _kpi(col):
        if port_table is None or port_table.empty:
            return 0
        tot_mask = port_table.iloc[:, 0].astype(str).str.startswith("📊")
        if tot_mask.any() and col in port_table.columns:
            return port_table[tot_mask][col].values[0]
        return port_table[col].sum() if col in port_table.columns else 0

    tot_debt  = _kpi("إجمالي المديونية")
    tot_coll  = _kpi("إجمالي التحصيل")
    tot_cust  = _kpi("عدد العملاء")
    coll_pct  = (tot_coll / tot_debt * 100) if tot_debt > 0 else 0
    cnt_cont  = _kpi("تم التوصل")
    cnt_noans = _kpi("لا يرد ومغلق")
    cnt_rate  = _kpi("نسبة تم التوصل %")

    kpi_defs = [
        ("👥 إجمالي العملاء",     f"{tot_cust:,.0f}",    1,  2),
        ("💰 إجمالي المديونية",   f"{tot_debt:,.0f} ﷼",  3,  4),
        ("💵 إجمالي التحصيل",    f"{tot_coll:,.0f} ﷼",  5,  6),
        ("📈 نسبة التحصيل",      f"{coll_pct:.1f}%",     7,  8),
        ("📞 تم التوصل",          f"{cnt_cont:,.0f}",     9,  10),
        ("📵 لا يرد ومغلق",      f"{cnt_noans:,.0f}",    11, 12),
        ("📈 نسبة التوصل",       f"{cnt_rate:.1f}%",     13, 14),
    ]

    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 10  # spacer

    for label, val_str, sc, ec in kpi_defs:
        ws.merge_cells(start_row=3, start_column=sc, end_row=3, end_column=ec)
        lc = ws.cell(row=3, column=sc, value=label)
        lc.fill = FILL_CARD
        lc.font = KPI_LBL_FONT
        lc.alignment = ALIGN_CENTER
        for r_c in ws.iter_rows(min_row=3, max_row=3, min_col=sc, max_col=ec):
            for c in r_c:
                c.border = BORDER_KPI

        ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)
        vc = ws.cell(row=4, column=sc, value=val_str)
        vc.fill = FILL_CARD
        vc.font = KPI_VAL_FONT
        vc.alignment = ALIGN_CENTER
        for r_c in ws.iter_rows(min_row=4, max_row=4, min_col=sc, max_col=ec):
            for c in r_c:
                c.border = BORDER_KPI

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 3. TABLE 1: ملخص الأداء الشامل ونسب التوصل
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ROW_T1_TITLE = 6
    section_title(ws, ROW_T1_TITLE, f"📋  الجدول الأول: ملخص الأداء والتحصيل ونسب التوصل — حسب {first_col}", TOTAL_SPAN)

    headers_main = list(port_table.columns)
    ROW_T1_HDR = ROW_T1_TITLE + 1
    write_header_row(ws, ROW_T1_HDR, headers_main, fill=FILL_HDR, height=22)

    ROW_T1_DATA = ROW_T1_HDR + 1
    port_reset = port_table.reset_index(drop=True)
    for i, (_, row_data) in enumerate(port_reset.iterrows()):
        write_data_row(ws, ROW_T1_DATA + i, row_data, headers_main, alt=(i % 2 == 1))
        ws.row_dimensions[ROW_T1_DATA + i].height = 20
    T1_END = ROW_T1_DATA + len(port_reset) - 1

    # ── شارتس الجدول الأول ──
    CHART1_ROW = T1_END + 2
    n_pts = len(port_reset) - 1  # بدون الإجمالي
    if n_pts > 0 and "إجمالي التحصيل" in headers_main:
        ci_coll = headers_main.index("إجمالي التحصيل") + 1
        dr_coll = Reference(ws, min_col=ci_coll, min_row=ROW_T1_HDR, max_row=ROW_T1_HDR + n_pts)
        cr_cats = Reference(ws, min_col=1, min_row=ROW_T1_DATA, max_row=ROW_T1_DATA + n_pts - 1)
        add_styled_bar_chart(ws, f"A{CHART1_ROW}", f"💰 إجمالي التحصيل — حسب {first_col}", dr_coll, cr_cats, width=17, height=10, orient="bar")

    if n_pts > 0 and "نسبة التحصيل %" in headers_main:
        ci_rate = headers_main.index("نسبة التحصيل %") + 1
        dr_rate = Reference(ws, min_col=ci_rate, min_row=ROW_T1_HDR, max_row=ROW_T1_HDR + n_pts)
        add_styled_bar_chart(ws, f"I{CHART1_ROW}", f"📈 نسبة التحصيل % — حسب {first_col}", dr_rate, cr_cats, width=16, height=10, orient="col")

    # مساحة نزول للشارتس
    ROW_AFTER_C1 = CHART1_ROW + 20

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 4. TABLE 2: جدول عمر الدين وسنة فصل الخدمة
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    if vintage_table is not None and not vintage_table.empty:
        ROW_T2_TITLE = ROW_AFTER_C1
        section_title(ws, ROW_T2_TITLE, "⏳  الجدول الثاني: تحليل عمر الدين — المديونية والتحصيل ونسب الإنجاز حسب سنة فصل الخدمة", TOTAL_SPAN)

        headers_vin = list(vintage_table.columns)
        ROW_T2_HDR = ROW_T2_TITLE + 1
        write_header_row(ws, ROW_T2_HDR, headers_vin, fill=FILL_HDR_VIN, height=22)

        ROW_T2_DATA = ROW_T2_HDR + 1
        vin_reset = vintage_table.reset_index(drop=True)
        for i, (_, row_data) in enumerate(vin_reset.iterrows()):
            write_data_row(ws, ROW_T2_DATA + i, row_data, headers_vin, alt=(i % 2 == 1))
            ws.row_dimensions[ROW_T2_DATA + i].height = 20
        T2_END = ROW_T2_DATA + len(vin_reset) - 1

        # ── شارتس عمر الدين ──
        CHART2_ROW = T2_END + 2
        nv_pts = len(vin_reset) - 1
        if nv_pts > 0 and "إجمالي التحصيل" in headers_vin:
            ci_vcoll = headers_vin.index("إجمالي التحصيل") + 1
            dr_vcoll = Reference(ws, min_col=ci_vcoll, min_row=ROW_T2_HDR, max_row=ROW_T2_HDR + nv_pts)
            cr_vcats = Reference(ws, min_col=1, min_row=ROW_T2_DATA, max_row=ROW_T2_DATA + nv_pts - 1)
            add_styled_bar_chart(ws, f"A{CHART2_ROW}", "⏳ إجمالي التحصيل — حسب سنة فصل الخدمة", dr_vcoll, cr_vcats, width=17, height=10, orient="bar")

        # ── مساحة نزول لشارتس عمر الدين ──
        ROW_AFTER_C2 = CHART2_ROW + 20
    else:
        ROW_AFTER_C2 = ROW_AFTER_C1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 5. TABLE 3: أفضل المشرفين وأفضل 5 محصلين أداءً
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ROW_T3_TITLE = ROW_AFTER_C2
    section_title(ws, ROW_T3_TITLE, "🏆  الجدول الثالث: ترتيب أفضل 5 محصلين (Top 5 Collectors) وأفضل المشرفين أداءً في التحصيل", TOTAL_SPAN)

    ROW_T3_HDR = ROW_T3_TITLE + 1

    # رؤوس جدول المحصلين (أعمدة 1 - 4)
    col_hdrs = ["الترتيب", "المحصل", "إجمالي التحصيل", "المعدل %"]
    for i, h in enumerate(col_hdrs, 1):
        c = ws.cell(row=ROW_T3_HDR, column=i, value=h)
        c.fill = FILL_HDR
        c.font = HDR_FONT
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # رؤوس جدول المشرفين (أعمدة 6 - 9)
    sup_hdrs = ["الترتيب", "المشرف", "إجمالي التحصيل", "المعدل %"]
    for i, h in enumerate(sup_hdrs, 6):
        c = ws.cell(row=ROW_T3_HDR, column=i, value=h)
        c.fill = PatternFill("solid", fgColor="1E40AF")
        c.font = HDR_FONT
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    ws.row_dimensions[ROW_T3_HDR].height = 22

    # كتابة بيانات المحصلين والمشرفين جنبًا إلى جنب
    top5_cols = col_df.head(5).reset_index(drop=True) if col_df is not None and not col_df.empty else pd.DataFrame()
    top5_sups = sup_df.head(5).reset_index(drop=True) if sup_df is not None and not sup_df.empty else pd.DataFrame()
    medals = {1: "🥇 الأول", 2: "🥈 الثاني", 3: "🥉 الثالث", 4: "🏅 الرابع", 5: "🏅 الخامس"}

    max_t3_rows = max(len(top5_cols), len(top5_sups), 1)

    for r_i in range(max_t3_rows):
        curr_row = ROW_T3_HDR + 1 + r_i
        ws.row_dimensions[curr_row].height = 20
        alt = (r_i % 2 == 1)

        # ── المحصلين ──
        if r_i < len(top5_cols):
            c_data = top5_cols.iloc[r_i]
            rank_txt = medals.get(r_i + 1, f"#{r_i+1}")
            c_name = str(c_data.get('المحصل', ''))
            c_coll = float(c_data.get('إجمالي التحصيل', 0))
            c_rate = float(c_data.get('المعدل %', 0)) / 100.0

            c1 = ws.cell(row=curr_row, column=1, value=rank_txt)
            c2 = ws.cell(row=curr_row, column=2, value=c_name)
            c3 = ws.cell(row=curr_row, column=3, value=c_coll)
            c4 = ws.cell(row=curr_row, column=4, value=c_rate)

            c1.alignment = ALIGN_CENTER
            c2.alignment = ALIGN_RIGHT
            c3.alignment = ALIGN_CENTER
            c4.alignment = ALIGN_CENTER

            c3.number_format = '#,##0 "﷼"'
            c4.number_format = '0.0%'

            for ci in [c1, c2, c3, c4]:
                ci.fill = FILL_ALT if alt else FILL_WHITE
                ci.font = DATA_FONT
                ci.border = BORDER_THIN

        # ── المشرفين ──
        if r_i < len(top5_sups):
            s_data = top5_sups.iloc[r_i]
            rank_txt = medals.get(r_i + 1, f"#{r_i+1}")
            s_name = str(s_data.get('المشرف', ''))
            s_coll = float(s_data.get('إجمالي التحصيل', 0))
            s_rate = float(s_data.get('المعدل %', 0)) / 100.0

            s1 = ws.cell(row=curr_row, column=6, value=rank_txt)
            s2 = ws.cell(row=curr_row, column=7, value=s_name)
            s3 = ws.cell(row=curr_row, column=8, value=s_coll)
            s4 = ws.cell(row=curr_row, column=9, value=s_rate)

            s1.alignment = ALIGN_CENTER
            s2.alignment = ALIGN_RIGHT
            s3.alignment = ALIGN_CENTER
            s4.alignment = ALIGN_CENTER

            s3.number_format = '#,##0 "﷼"'
            s4.number_format = '0.0%'

            for si in [s1, s2, s3, s4]:
                si.fill = FILL_ALT if alt else FILL_WHITE
                si.font = DATA_FONT
                si.border = BORDER_THIN

    # ── شارت أفضل المحصلين إلى جانب الجدول ──
    if len(top5_cols) > 0:
        try:
            dr_col_pts = Reference(ws, min_col=3, min_row=ROW_T3_HDR, max_row=ROW_T3_HDR + len(top5_cols))
            cr_col_cats = Reference(ws, min_col=2, min_row=ROW_T3_HDR + 1, max_row=ROW_T3_HDR + len(top5_cols))
            add_styled_bar_chart(ws, f"K{ROW_T3_HDR}", "⭐ أفضل 5 محصلين أداءً في التحصيل", dr_col_pts, cr_col_cats, width=15, height=9, orient="bar")
        except Exception:
            pass

    # ضبط العرض التلقائي للأعمدة
    auto_width(ws, max_col=TOTAL_SPAN + 2)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 5. SHEET 2: تفاصيل السدادات
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws2 = wb.create_sheet(title="تفاصيل السدادات")
    ws2.views.sheetView[0].rightToLeft = True
    ws2.views.sheetView[0].showGridLines = True

    df_pay_exp = df_pay_filtered.drop(columns=["_pay_date", "m_amt"], errors="ignore")
    pay_hdrs   = list(df_pay_exp.columns)

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(pay_hdrs))
    pb = ws2.cell(row=1, column=1, value=f"💳  سجل السدادات التفصيلية — {report_date}")
    pb.fill = FILL_BANNER
    pb.font = BANNER_FONT
    pb.alignment = ALIGN_CENTER
    ws2.row_dimensions[1].height = 28

    write_header_row(ws2, 2, pay_hdrs, fill=FILL_HDR, height=20)
    for r_i, row in df_pay_exp.iterrows():
        row_n = r_i + 3
        alt = (r_i % 2 == 1)
        for c_i, val in enumerate(row, 1):
            c = ws2.cell(row=row_n, column=c_i, value=str(val) if pd.notna(val) else "")
            c.font = DATA_FONT
            c.fill = FILL_ALT if alt else FILL_WHITE
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER
        ws2.row_dimensions[row_n].height = 18

    auto_width(ws2, max_col=len(pay_hdrs) + 1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
